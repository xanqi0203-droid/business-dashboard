#!/usr/bin/env python3
"""
BEAM — 热点话题数据 KPI 提取脚本
从热点数据文件中提取关键指标

支持格式：
  - Excel (.xlsx, .xls)
  - CSV (.csv)
  - JSON (.json)

输出：
  {
    "total_topics": 话题总数,
    "hot_topics": 热门话题数（热度>阈值）,
    "total_engagement": 总互动量,
    "avg_heat_score": 平均热度分数,
    "topics_growth": 话题增长率（%）,
    "engagement_growth": 互动增长率（%）,
    "top_topics": [
      {"name": "话题名", "heat_score": 热度, "engagement": 互动量, "platform": "平台"},
      ...
    ],
    "platform_distribution": {"微博": 数量, "抖音": 数量, ...}
  }

用法：
  python extract_hotspot_kpis.py <数据文件> [--output kpis.json]
"""

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict, List


def load_data(file_path: str) -> List[Dict[str, Any]]:
    """加载数据文件，返回标准化的记录列表"""
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"数据文件不存在: {file_path}")

    suffix = path.suffix.lower()

    # JSON 格式
    if suffix == '.json':
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, list):
                return data
            elif isinstance(data, dict) and 'data' in data:
                return data['data']
            else:
                raise ValueError("JSON 格式错误：需要数组或包含 'data' 字段的对象")

    # CSV 格式
    elif suffix == '.csv':
        import csv
        with open(path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            return list(reader)

    # Excel 格式
    elif suffix in ['.xlsx', '.xls']:
        try:
            import openpyxl
        except ImportError:
            print("ERROR: 需要安装 openpyxl: pip install openpyxl", file=sys.stderr)
            sys.exit(1)

        wb = openpyxl.load_workbook(path, data_only=True)

        # 尝试找到热点数据表（按关键词匹配）
        sheet = None
        for name in wb.sheetnames:
            if any(keyword in name for keyword in ['热点', '话题', '舆情', 'hot', 'topic']):
                sheet = wb[name]
                break

        if sheet is None:
            sheet = wb.active

        # 读取表头和数据
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            if not any(row):  # 跳过空行
                continue
            record = {headers[i]: row[i] for i in range(len(headers))}
            data.append(record)

        return data

    else:
        raise ValueError(f"不支持的文件格式: {suffix}")


def normalize_field(record: Dict[str, Any], possible_names: List[str], default=None):
    """从记录中查找字段（支持多种可能的字段名）"""
    for name in possible_names:
        for key in record.keys():
            if name.lower() in key.lower():
                value = record[key]
                # 处理空值
                if value is None or value == '':
                    return default
                # 尝试转换为数字
                if isinstance(value, str):
                    value = value.replace(',', '').replace('万', '0000')
                    try:
                        return float(value)
                    except ValueError:
                        return value
                return value
    return default


def classify_topic(topic_name: str) -> str:
    """根据关键词将话题分类"""
    topic_lower = topic_name.lower()

    # AI 相关关键词
    ai_keywords = ['ai', '人工智能', '大模型', 'gpt', 'chatgpt', '机器学习', '深度学习',
                   '神经网络', '算法', '智能', 'llm', 'claude', 'gemini']

    # 科技相关关键词
    tech_keywords = ['科技', '芯片', '半导体', '5g', '6g', '量子', '新能源', '电动车',
                     '特斯拉', '华为', '小米', '苹果', 'iphone', '手机', '电脑',
                     '互联网', '区块链', '元宇宙', 'vr', 'ar']

    # 银行/金融相关关键词
    bank_keywords = ['银行', '金融', '理财', '贷款', '存款', '利率', '股市', '基金',
                     '证券', '保险', '支付', '央行', '货币', '经济', '投资']

    # 娱乐相关关键词
    entertainment_keywords = ['娱乐', '明星', '电影', '电视剧', '综艺', '音乐', '演唱会',
                             '票房', '导演', '演员', '歌手', '偶像', '网红', '直播']

    # 按优先级匹配
    for keyword in ai_keywords:
        if keyword in topic_lower:
            return "AI"

    for keyword in tech_keywords:
        if keyword in topic_lower:
            return "科技"

    for keyword in bank_keywords:
        if keyword in topic_lower:
            return "银行"

    for keyword in entertainment_keywords:
        if keyword in topic_lower:
            return "娱乐"

    return None  # 不属于目标分类


def compute_kpis(data: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    计算 KPI 指标

    需要根据实际数据字段名调整 normalize_field 的参数
    常见字段名映射：
      - 话题名称: topic, title, name, 话题, 标题
      - 热度分数: heat, score, hot, 热度, 分数
      - 互动量: engagement, interaction, 互动, 评论, 点赞
      - 平台: platform, source, 来源, 平台
      - 时间: time, date, created, 时间, 日期
    """

    if not data:
        return {
            "total_topics": 0,
            "hot_topics": 0,
            "total_engagement": 0,
            "avg_heat_score": 0,
            "topics_growth": 0,
            "engagement_growth": 0,
            "top_topics": [],
            "platform_distribution": {}
        }

    # 提取标准化字段
    topics = []
    for record in data:
        topic = {
            "name": normalize_field(record, ["topic", "title", "name", "话题", "标题"], "未知话题"),
            "heat_score": normalize_field(record, ["heat", "score", "hot", "热度", "分数"], 0),
            "engagement": normalize_field(record, ["engagement", "interaction", "互动", "评论", "点赞"], 0),
            "platform": normalize_field(record, ["platform", "source", "来源", "平台"], "未知"),
            "url": normalize_field(record, ["url", "link", "href", "链接", "source_url"], ""),
        }
        # 自动分类
        topic["category"] = classify_topic(str(topic["name"]))
        topics.append(topic)

    # 基础指标
    total_topics = len(topics)
    hot_threshold = 1000
    hot_topics = sum(1 for t in topics if t["heat_score"] >= hot_threshold)

    total_engagement = sum(t["engagement"] for t in topics)
    avg_heat_score = sum(t["heat_score"] for t in topics) / total_topics if total_topics > 0 else 0

    # 按分类分组（只保留目标分类），每类按热度排序取前3
    categories = ["AI", "科技", "银行", "娱乐"]
    topics_by_category = {}
    for cat in categories:
        cat_topics = [t for t in topics if t["category"] == cat]
        cat_topics.sort(key=lambda x: x["heat_score"], reverse=True)
        topics_by_category[cat] = [
            {"name": t["name"], "heat_score": t["heat_score"],
             "engagement": int(t["engagement"]), "platform": t["platform"],
             "url": t.get("url", "")}
            for t in cat_topics[:3]
        ]

    # Top 话题（所有分类合并排序）
    top_topics = sorted(topics, key=lambda x: x["heat_score"], reverse=True)[:10]

    # 平台分布
    platform_distribution = {}
    for topic in topics:
        platform = topic["platform"]
        platform_distribution[platform] = platform_distribution.get(platform, 0) + 1

    # 增长率（需要历史数据对比，这里暂时返回 0）
    # 实际使用时可以读取昨天的 KPI 文件进行对比
    topics_growth = 0
    engagement_growth = 0

    return {
        "total_topics": total_topics,
        "hot_topics": hot_topics,
        "total_engagement": int(total_engagement),
        "avg_heat_score": round(avg_heat_score, 2),
        "topics_growth": 0,
        "engagement_growth": 0,
        "topics_by_category": topics_by_category,
        "top_topics": [
            {"name": t["name"], "heat_score": t["heat_score"],
             "engagement": int(t["engagement"]), "platform": t["platform"],
             "url": t.get("url", "")}
            for t in top_topics
        ],
        "platform_distribution": platform_distribution
    }


def main():
    parser = argparse.ArgumentParser(description="提取热点数据 KPI")
    parser.add_argument("data_file", help="数据文件路径")
    parser.add_argument("--output", default=".tmp/kpis.json", help="输出 JSON 文件路径")

    args = parser.parse_args()

    try:
        # 加载数据
        print(f"加载数据: {args.data_file}")
        data = load_data(args.data_file)
        print(f"✓ 读取 {len(data)} 条记录")

        # 计算 KPI
        print("计算 KPI...")
        kpis = compute_kpis(data)
        print(f"✓ 话题总数: {kpis['total_topics']}")
        print(f"✓ 热门话题: {kpis['hot_topics']}")
        print(f"✓ 总互动量: {kpis['total_engagement']:,}")

        # 保存结果
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(kpis, f, ensure_ascii=False, indent=2)

        print(f"✓ KPI 已保存到: {output_path}")

    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
